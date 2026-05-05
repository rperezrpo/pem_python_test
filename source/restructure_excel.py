"""
Restructure the Excel for the new model:
- Remove: f_SL, f_LL, f_Ag (fraction lost per event - replaced by u_i)
- Remove: c_install_j, N_prot_j (move to per-element)
- Add per element: u_i (units damaged per event)
- Add per element: s_i (share of attacks targeting this element)
- Add per element: c_install_j_i (install cost per unit)
- Add per element: N_prot_j_i (units protected)
- Add: T (time horizon for the analysis, in years)
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

PATH = Path("/mnt/user-data/outputs/bear_attack_model_variables.xlsx")

# Build the new variable list from scratch (cleanest approach)
ROWS = [
    # ---------- Bear pressure ----------
    ("Average bear incursions per year",
     "Estimated number of bear incursions per year per confident bear in the broader area surrounding the property - a baseline pressure metric used to scale the attack rate.",
     "I_y", 2200),
    ("Number of confident bears",
     "Number of bears in the broader area that are habituated/confident enough to approach human settlements and cause damage. Multiplied by I_y to scale the property-level encounter rate.",
     "N_bears", 5),

    # ---------- Environment modifiers (now feed into λ, not per-element formulas) ----------
    ("Food availability modifier",
     "Multiplicative adjustment on the property-level attack rate based on wild food availability. Low mast/berry years (>1) increase the attack rate; abundant wild food (<1) decreases it.",
     "P(D_i|food)", 1.20),

    # ---------- Per-element damage probabilities (still used to combine with encounter share) ----------
    ("Damage to small livestock",
     "Probability of damage to small livestock (sheep, goats, poultry) given bear presence and encounter. Combined with the encounter share to determine how many events target this element type.",
     "P(D_SL|B,E)", 0.35),
    ("Damage to large livestock",
     "Probability of damage to large livestock (cattle, horses) given bear presence and encounter.",
     "P(D_LL|B,E)", 0.05),
    ("Damage to agriculture",
     "Probability of damage to agricultural assets (orchards, beehives, crops) given bear presence and encounter.",
     "P(D_Ag|B,E)", 0.45),

    # ---------- Encounter share per element (NEW: must sum to 100%) ----------
    ("Encounter share - small livestock",
     "Share of all bear encounters at this property that target small livestock. The three element shares must sum to 100%.",
     "s_enc_SL", 0.40),
    ("Encounter share - large livestock",
     "Share of all bear encounters at this property that target large livestock.",
     "s_enc_LL", 0.10),
    ("Encounter share - agriculture",
     "Share of all bear encounters at this property that target agricultural assets.",
     "s_enc_Ag", 0.50),

    # ---------- Maintenance probability ----------
    ("Maintenance of protective measure",
     "Probability that the protective measure is properly maintained (functional) at the time of an encounter. A maintained measure is assumed to be 100% effective.",
     "P(M_j)", 0.85),
    ("Failure of protective measure",
     "Probability that the protective measure fails (1 - P(M_j)). Auto-derived from P(M_j).",
     "P(¬M_j)", 0.15),

    # ---------- Per-element: units exposed ----------
    ("Number of small livestock units",
     "Total number of small livestock units (head) present and exposed on the property.",
     "N_SL", 50),
    ("Number of large livestock units",
     "Total number of large livestock units (head) present and exposed on the property.",
     "N_LL", 10),
    ("Number of agricultural units",
     "Total number of agricultural units exposed on the property (e.g., beehives, fruit trees).",
     "N_Ag", 20),

    # ---------- Per-element: units typically damaged per event (NEW: replaces f_i) ----------
    ("Units damaged per event - small livestock",
     "Expected number of small livestock units lost or damaged in a single bear damage event when the bear has access. A behavioral parameter of bear-on-flock attacks; bears typically take 3-5 head per event.",
     "u_SL", 4),
    ("Units damaged per event - large livestock",
     "Expected number of large livestock units (typically calves) lost in a single bear damage event when the bear has access. Bears usually take 1 calf per event.",
     "u_LL", 1),
    ("Units damaged per event - agriculture",
     "Expected number of agricultural units (e.g., beehives) destroyed in a single bear damage event when the bear has access. Bears tend to destroy multiple beehives per visit.",
     "u_Ag", 6),

    # ---------- Per-element: compensation costs ----------
    ("Compensation cost per unit - small livestock",
     "Compensation paid per small livestock unit lost or damaged.",
     "c_SL", 200),
    ("Compensation cost per unit - large livestock",
     "Compensation paid per large livestock unit lost or damaged.",
     "c_LL", 1500),
    ("Compensation cost per unit - agriculture",
     "Compensation paid per agricultural unit lost or damaged.",
     "c_Ag", 400),

    # ---------- Per-element: protection ----------
    ("Number of units protected - small livestock",
     "Number of small livestock units covered by the protective measure (e.g., behind an electric fence or in a guarded pen).",
     "N_prot_SL", 40),
    ("Number of units protected - large livestock",
     "Number of large livestock units covered by the protective measure.",
     "N_prot_LL", 8),
    ("Number of units protected - agriculture",
     "Number of agricultural units covered by the protective measure (e.g., beehives behind a bear-proof fence).",
     "N_prot_Ag", 18),

    # ---------- Per-element: installation cost (NEW: was global) ----------
    ("Installation cost per unit - small livestock",
     "One-time installation cost of the protective measure for small livestock, per unit protected (e.g., per sheep place in a guarded enclosure).",
     "c_install_SL", 60),
    ("Installation cost per unit - large livestock",
     "One-time installation cost of the protective measure for large livestock, per unit protected.",
     "c_install_LL", 200),
    ("Installation cost per unit - agriculture",
     "One-time installation cost of the protective measure for agriculture, per unit protected (e.g., per beehive on a fenced platform).",
     "c_install_Ag", 80),

    # ---------- Per-element: maintenance cost (NEW: was global) ----------
    ("Maintenance cost per unit protected - small livestock",
     "Recurring annual maintenance cost of the protective measure for small livestock, per unit protected (e.g., per sheep place).",
     "c_maintenance_SL", 5),
    ("Maintenance cost per unit protected - large livestock",
     "Recurring annual maintenance cost of the protective measure for large livestock, per unit protected.",
     "c_maintenance_LL", 20),
    ("Maintenance cost per unit protected - agriculture",
     "Recurring annual maintenance cost of the protective measure for agriculture, per unit protected (e.g., per beehive on a fenced platform).",
     "c_maintenance_Ag", 5),

    # ---------- Global protective-measure parameters ----------
    ("Lifespan of protective measure",
     "Expected useful lifespan of the protective measure, in years. After this period a re-investment would be needed.",
     "lifespan_j", 10),

    # ---------- Time horizon (NEW) ----------
    ("Analysis time horizon",
     "Number of years over which the cumulative cost-benefit is evaluated. Installation costs are paid in year 0; maintenance and damages accrue each year.",
     "T", 5),
]

wb = Workbook()
ws = wb.active
ws.title = "Model Variables"
ws.append(["Name", "Description", "Notation", "Value"])
for r in ROWS:
    ws.append(list(r))

# Formatting
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color="305496")
header_align = Alignment(horizontal="left", vertical="center")
body_font = Font(name="Arial", size=11)
body_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx in range(1, 5):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

yellow_fill = PatternFill("solid", start_color="FFF2CC")
for r in range(2, ws.max_row + 1):
    for c in range(1, 5):
        cell = ws.cell(row=r, column=c)
        cell.font = body_font
        cell.alignment = body_align
        cell.border = border
    ws.cell(row=r, column=4).fill = yellow_fill

ws.column_dimensions["A"].width = 44
ws.column_dimensions["B"].width = 80
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 14
ws.row_dimensions[1].height = 22
for r in range(2, ws.max_row + 1):
    ws.row_dimensions[r].height = 45
ws.freeze_panes = "A2"

wb.save(PATH)
print(f"Saved {len(ROWS)} variables to {PATH}")
