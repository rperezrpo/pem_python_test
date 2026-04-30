
# py source/pipeline.py

from pathlib import Path
from openpyxl import load_workbook

from create_html import build_html

EXCEL_PATH = Path("source/bear_attack_model_variables.xlsx")
HTML_PATH = Path("source/bear_cost_benefit.html")


def read_variables(path: Path):
    wb = load_workbook(path)
    ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if not name:
            continue
        rows.append({
            "name": name,
            "description": ws.cell(row=r, column=2).value,
            "notation": ws.cell(row=r, column=3).value,
            "value": ws.cell(row=r, column=4).value,
        })
    return rows


def main():
    print("→ Reading variables from Excel...")
    variables = read_variables(EXCEL_PATH)
    print(f"  loaded {len(variables)} variables")

    missing = [v["notation"] for v in variables if v["value"] is None]
    if missing:
        print(f"  ⚠ no value for: {missing}")

    print("→ Generating interactive HTML...")
    html = build_html(variables)
    HTML_PATH.write_text(html, encoding="utf-8")
    print(f"  wrote {HTML_PATH} ({len(html):,} bytes)")

    print("\n✓ Pipeline complete.")
    print(f"  Excel: {EXCEL_PATH}")
    print(f"  HTML:  {HTML_PATH}")


if __name__ == "__main__":
    main()
